import io
import pandas as pd
from datetime import timedelta, date
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================
# CONFIG
# ==============================

st.set_page_config(
    page_title="Calculadora de Horas Ponto",
    page_icon="⏱️",
    layout="wide"
)

# ==============================
# CSS
# ==============================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

html, body, [data-testid="stAppViewContainer"] {
    background: #0f1117 !important;
    font-family: 'DM Sans', sans-serif;
}

[data-testid="stAppViewContainer"] {
    background: linear-gradient(135deg, #0f1117 0%, #1a1d2e 50%, #0f1117 100%) !important;
}

[data-testid="stHeader"] { background: transparent !important; }
[data-testid="stSidebar"] { display: none; }

.block-container {
    padding: 2rem 3rem !important;
    max-width: 1100px !important;
}

/* Hero */
.hero {
    text-align: center;
    padding: 3rem 0 2rem;
    position: relative;
}
.hero-badge {
    display: inline-block;
    background: rgba(0, 212, 170, 0.12);
    border: 1px solid rgba(0, 212, 170, 0.3);
    color: #00d4aa;
    font-family: 'DM Mono', monospace;
    font-size: 0.72rem;
    letter-spacing: 0.15em;
    padding: 0.4rem 1rem;
    border-radius: 100px;
    margin-bottom: 1.5rem;
    text-transform: uppercase;
}
.hero-title {
    font-size: clamp(2.2rem, 5vw, 3.4rem);
    font-weight: 700;
    color: #ffffff;
    line-height: 1.1;
    margin-bottom: 0.8rem;
    letter-spacing: -0.02em;
}
.hero-title span {
    background: linear-gradient(135deg, #00d4aa, #0099ff);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.hero-sub {
    color: #6b7280;
    font-size: 1.05rem;
    font-weight: 400;
    max-width: 480px;
    margin: 0 auto;
    line-height: 1.6;
}

/* Divider */
.divider {
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(255,255,255,0.08), transparent);
    margin: 2rem 0;
}

/* Upload box */
.upload-label {
    color: #9ca3af;
    font-size: 0.85rem;
    font-weight: 500;
    letter-spacing: 0.05em;
    text-transform: uppercase;
    margin-bottom: 0.5rem;
    display: block;
}

[data-testid="stFileUploader"] {
    background: rgba(255,255,255,0.03) !important;
    border: 1.5px dashed rgba(0, 212, 170, 0.3) !important;
    border-radius: 16px !important;
    padding: 1.5rem !important;
    transition: all 0.3s ease;
}
[data-testid="stFileUploader"]:hover {
    border-color: rgba(0, 212, 170, 0.6) !important;
    background: rgba(0, 212, 170, 0.04) !important;
}
[data-testid="stFileUploader"] label {
    color: #9ca3af !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stFileUploaderDropzoneInstructions"] {
    color: #6b7280 !important;
}

/* Metrics */
.metrics-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin: 1.5rem 0;
}
.metric-card {
    background: rgba(255,255,255,0.04);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 16px;
    padding: 1.4rem 1.2rem;
    text-align: center;
    transition: all 0.3s ease;
    position: relative;
    overflow: hidden;
}
.metric-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, #00d4aa, #0099ff);
    opacity: 0;
    transition: opacity 0.3s;
}
.metric-card:hover::before { opacity: 1; }
.metric-card:hover {
    background: rgba(255,255,255,0.07);
    border-color: rgba(0, 212, 170, 0.2);
    transform: translateY(-2px);
}
.metric-icon { font-size: 1.5rem; margin-bottom: 0.5rem; }
.metric-value {
    font-size: 2rem;
    font-weight: 700;
    color: #ffffff;
    font-family: 'DM Mono', monospace;
    line-height: 1;
    margin-bottom: 0.3rem;
}
.metric-value.accent { color: #00d4aa; }
.metric-label {
    color: #6b7280;
    font-size: 0.78rem;
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 0.08em;
}

/* Section title */
.section-title {
    color: #e5e7eb;
    font-size: 1rem;
    font-weight: 600;
    letter-spacing: -0.01em;
    margin-bottom: 0.8rem;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}
.section-title::after {
    content: '';
    flex: 1;
    height: 1px;
    background: rgba(255,255,255,0.08);
}

/* Dataframe override */
[data-testid="stDataFrame"] {
    border-radius: 12px !important;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.08) !important;
}

/* Selectbox */
[data-testid="stSelectbox"] > div > div {
    background: rgba(255,255,255,0.05) !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
    border-radius: 10px !important;
    color: #e5e7eb !important;
}

/* Download button */
[data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #00d4aa, #0099ff) !important;
    color: #0f1117 !important;
    font-weight: 700 !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 0.75rem 2rem !important;
    font-size: 0.95rem !important;
    letter-spacing: 0.01em !important;
    width: 100%;
    transition: all 0.3s ease !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stDownloadButton"] > button:hover {
    opacity: 0.9 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 8px 24px rgba(0, 212, 170, 0.3) !important;
}

/* Spinner */
[data-testid="stSpinner"] { color: #00d4aa !important; }

/* Warning/error */
[data-testid="stAlert"] {
    border-radius: 12px !important;
    border: none !important;
    font-family: 'DM Sans', sans-serif !important;
}

/* Expander */
[data-testid="stExpander"] {
    background: rgba(255,255,255,0.03) !important;
    border: 1px solid rgba(255,255,255,0.08) !important;
    border-radius: 12px !important;
}

/* Footer */
.footer {
    text-align: center;
    color: #374151;
    font-size: 0.78rem;
    padding: 2rem 0 1rem;
    font-family: 'DM Mono', monospace;
}

/* Columns fix */
[data-testid="column"] { padding: 0 0.4rem !important; }

/* Hide streamlit branding */
#MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ==============================
# FUNÇÕES
# ==============================

def feriados_brasil(ano):
    return {
        date(ano, 1, 1), date(ano, 4, 21), date(ano, 5, 1),
        date(ano, 9, 7), date(ano, 10, 12), date(ano, 11, 2),
        date(ano, 11, 15), date(ano, 12, 25),
    }

def eh_fds_ou_feriado(d):
    return d.weekday() >= 5 or d in feriados_brasil(d.year)

def normalizar_colunas(df):
    df.columns = (
        df.columns.str.strip().str.upper()
        .str.normalize("NFD").str.encode("ascii", "ignore").str.decode("ascii")
    )
    return df

def detectar_cabecalho(df_raw):
    for i in range(min(10, len(df_raw))):
        try:
            linha = (
                df_raw.iloc[i].dropna().astype(str).str.upper()
                .str.normalize("NFD").str.encode("ascii", "ignore").str.decode("ascii")
                .tolist()
            )
            if "DATA" in linha and "ENTRADA" in linha:
                return i
        except Exception:
            continue
    return None

def tratar_hora(valor, data_ref):
    try:
        return pd.to_datetime(f"{data_ref} {str(valor).strip()}")
    except Exception:
        return None

@st.cache_data(show_spinner=False)
def processar_planilha(conteudo: bytes):
    df_raw = pd.read_excel(io.BytesIO(conteudo), header=None)
    header_row = detectar_cabecalho(df_raw)
    if header_row is None:
        return pd.DataFrame(), ["Cabeçalho com 'DATA' e 'ENTRADA' não encontrado."]

    df = pd.read_excel(io.BytesIO(conteudo), header=header_row)
    df = normalizar_colunas(df)

    colunas_necessarias = {"DATA", "ENTRADA", "SAIDA", "NOME DO MEDICO"}
    faltando = colunas_necessarias - set(df.columns)
    if faltando:
        return pd.DataFrame(), [f"Colunas não encontradas: {', '.join(faltando)}"]

    dias_semana = ["Segunda","Terça","Quarta","Quinta","Sexta","Sábado","Domingo"]
    resultados, avisos = [], []

    for idx, row in df.iterrows():
        linha_num = idx + header_row + 2
        try:
            data = pd.to_datetime(row["DATA"]).date()
            entrada = tratar_hora(row["ENTRADA"], data)
            saida   = tratar_hora(row["SAIDA"],   data)
            if entrada is None or saida is None:
                avisos.append(f"Linha {linha_num}: hora inválida — ignorada.")
                continue
            if saida < entrada:
                saida += timedelta(days=1)
            horas         = (saida - entrada).total_seconds() / 3600
            extras_brutas = max(0.0, horas - 8.0)
            fds           = eh_fds_ou_feriado(data)
            extras_calc   = extras_brutas * (2.0 if fds else 1.5)
            resultados.append({
                "Funcionário":         str(row["NOME DO MEDICO"]).strip(),
                "Data":                data.strftime("%d/%m/%Y"),
                "Dia da Semana":       dias_semana[data.weekday()],
                "Feriado/FDS":         "Sim" if fds else "Não",
                "Entrada":             entrada.strftime("%H:%M"),
                "Saída":               saida.strftime("%H:%M"),
                "Horas Trabalhadas":   round(horas, 2),
                "Horas Extras (H)":    round(extras_brutas, 2),
                "Horas Extras (calc)": round(extras_calc, 2),
            })
        except Exception as e:
            avisos.append(f"Linha {linha_num}: erro ({e}) — ignorada.")

    return pd.DataFrame(resultados), avisos

def formatar_aba(ws):
    verde, branco, cinza = "006341", "FFFFFF", "F2F2F2"
    borda = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    for cell in ws[1]:
        cell.font = Font(bold=True, color=branco, name="Arial", size=11)
        cell.fill = PatternFill("solid", start_color=verde)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda
        ws.row_dimensions[1].height = 20
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = cinza if i % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center")
            cell.border = borda
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 42)

def gerar_excel(df):
    resumo = (
        df.groupby("Funcionário")[["Horas Trabalhadas","Horas Extras (H)","Horas Extras (calc)"]]
        .sum().round(2).reset_index()
    )
    resumo.columns = ["Funcionário","Total H. Trabalhadas","Total H. Extras (H)","Total H. Extras (calc)"]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        resumo.to_excel(writer, sheet_name="Resumo",    index=False)
        df.to_excel(    writer, sheet_name="Detalhado", index=False)

    wb = load_workbook(buf)
    verde, branco = "006341", "FFFFFF"
    borda = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        formatar_aba(ws)
        if nome_aba == "Resumo":
            ultima = ws.max_row + 1
            ws.cell(ultima, 1, "TOTAL").font = Font(bold=True, name="Arial", size=11, color=branco)
            ws.cell(ultima, 1).fill = PatternFill("solid", start_color=verde)
            ws.cell(ultima, 1).alignment = Alignment(horizontal="center")
            ws.cell(ultima, 1).border = borda
            for col in range(2, ws.max_column + 1):
                letra = get_column_letter(col)
                c = ws.cell(ultima, col)
                c.value = f"=SUM({letra}2:{letra}{ultima-1})"
                c.font = Font(bold=True, name="Arial", size=11, color=branco)
                c.fill = PatternFill("solid", start_color=verde)
                c.alignment = Alignment(horizontal="center")
                c.number_format = "0.00"
                c.border = borda

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), resumo

# ==============================
# INTERFACE
# ==============================

# Hero
st.markdown("""
<div class="hero">
    <div class="hero-badge">⏱ Prefeitura de Chapecó</div>
    <h1 class="hero-title">Calculadora de<br><span>Horas Ponto</span></h1>
    <p class="hero-sub">Envie a planilha de registros e obtenha o somatório de horas por médico em segundos.</p>
</div>
<div class="divider"></div>
""", unsafe_allow_html=True)

# Upload
st.markdown('<span class="upload-label">📂 Planilha de registros (.xlsx)</span>', unsafe_allow_html=True)
arquivo = st.file_uploader("", type=["xlsx"], label_visibility="collapsed")

if not arquivo:
    st.markdown("""
    <div style="margin-top:2rem; padding:1.5rem; background:rgba(255,255,255,0.02);
         border:1px solid rgba(255,255,255,0.06); border-radius:16px; color:#4b5563; font-size:0.9rem; line-height:1.8;">
        <strong style="color:#6b7280;">Colunas esperadas na planilha:</strong><br>
        <code style="color:#00d4aa; font-family:'DM Mono',monospace;">DATA</code> &nbsp;·&nbsp;
        <code style="color:#00d4aa; font-family:'DM Mono',monospace;">ENTRADA</code> &nbsp;·&nbsp;
        <code style="color:#00d4aa; font-family:'DM Mono',monospace;">SAIDA</code> &nbsp;·&nbsp;
        <code style="color:#00d4aa; font-family:'DM Mono',monospace;">NOME DO MEDICO</code>
    </div>
    """, unsafe_allow_html=True)
    st.markdown('<div class="footer">calculadora de horas ponto · prefeitura de chapecó</div>', unsafe_allow_html=True)
    st.stop()

conteudo = arquivo.read()

with st.spinner("Processando registros..."):
    df, avisos = processar_planilha(conteudo)

if avisos:
    with st.expander(f"⚠️ {len(avisos)} aviso(s)", expanded=False):
        for a in avisos:
            st.warning(a)

if df.empty:
    st.error("Nenhum dado válido encontrado na planilha.")
    st.stop()

excel_bytes, resumo = gerar_excel(df)

# Métricas
total_func    = len(resumo)
total_reg     = len(df)
total_horas   = round(df["Horas Trabalhadas"].sum(), 1)
total_extras  = round(df["Horas Extras (calc)"].sum(), 1)

st.markdown(f"""
<div class="metrics-grid">
    <div class="metric-card">
        <div class="metric-icon">👥</div>
        <div class="metric-value">{total_func}</div>
        <div class="metric-label">Funcionários</div>
    </div>
    <div class="metric-card">
        <div class="metric-icon">📋</div>
        <div class="metric-value">{total_reg}</div>
        <div class="metric-label">Registros</div>
    </div>
    <div class="metric-card">
        <div class="metric-icon">🕐</div>
        <div class="metric-value">{total_horas}</div>
        <div class="metric-label">Horas Totais</div>
    </div>
    <div class="metric-card">
        <div class="metric-icon">⚡</div>
        <div class="metric-value accent">{total_extras}</div>
        <div class="metric-label">Horas Extras (calc)</div>
    </div>
</div>
<div class="divider"></div>
""", unsafe_allow_html=True)

# Resumo
st.markdown('<div class="section-title">Resumo por Funcionário</div>', unsafe_allow_html=True)
st.dataframe(resumo, use_container_width=True, hide_index=True)

st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

# Detalhado
st.markdown('<div class="section-title">Detalhamento por Registro</div>', unsafe_allow_html=True)
funcionarios = ["Todos"] + sorted(df["Funcionário"].unique().tolist())
selecionado  = st.selectbox("Filtrar por funcionário:", funcionarios)
df_exibir    = df if selecionado == "Todos" else df[df["Funcionário"] == selecionado]
st.dataframe(df_exibir, use_container_width=True, hide_index=True)

st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

# Download
st.download_button(
    label="📥 Baixar Relatório Excel",
    data=excel_bytes,
    file_name="relatorio_horas.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.markdown('<div class="footer">calculadora de horas ponto · prefeitura de chapecó</div>', unsafe_allow_html=True)
